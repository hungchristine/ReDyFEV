# -*- coding: utf-8 -*-
"""
This script runs the data extraction, data cleaning, calculations and visualization
scripts used in "Regionalized footprints of battery electric vehicles in Europe"

Users can run two types of experiments concurrently; the electricity sampling period
('el_experiments'), and vehicle parameters ('BEV_experiments')
"""
import logging
from datetime import datetime, timezone, timedelta
import pytz
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import numpy as np
import itertools

import bentso_extract
import entso_data_clean
import hybridized_impact_factors as hybrid
import BEV_footprints_calculations as calc
import BEV_visualization as viz

logname = 'run ' + datetime.now().strftime('%d-%m-%y, %H-%M') + '.log'
logging.basicConfig(filename=logname,
                            format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                            datefmt='%H:%M:%S',
                            level=logging.INFO)


if __name__ == "__main__":

    # Define experiment running parameters
    new_query = 0  # trigger for re-querying ENTSO-E database (takes time!)
    calc_hybrid = 1 # trigger for re-calculating hybridized impact factors (takes time!); 'fresh start' runs need value 1
    segments = ['A', 'C', 'JC', 'JE', 'mini', 'medium', 'compact SUV', 'mid-size SUV']  # valid entries for segment
    segment_dict = {'mini': 'A', 'medium':'C', 'compact SUV':'JC', 'mid-size SUV':'JE'} # long-form to short form segments

    # definition of minimum data for each experiment type
    exp_type_dict = {'fullyear':set(['year']),
                     'subyear': set(['year', 'start', 'end']),
                     'country_fp': set(['year', 'start', 'country', 'segment'])
                     }

    # define time period for experiment sampling (for ENTSO-E querying)
    # provide start and end period for subannual analysis, or just start for a country footprint
    start = datetime(2021, 5, 14, 22, 44, tzinfo=pytz.timezone('Europe/Berlin'))  # year, month, day, hour, minute, timezone
    end = datetime(2021, 5, 14, 23, 00, tzinfo=pytz.timezone('Europe/Berlin'))

    if start.year != end.year:
        logging.error('Cannot do cross-year comparisons!')

    # define electricity experiments (dict of dicts: {experiment_name: {experiment_params}})
    # accepted forms for experiment_params in el experiments (examples below):
    # -- 'year': single-element list with the full year of analysis
    # -- 3-item dict with year, start and stop periods as datetime objects
    # -- 4-item dict with year, time of assessment (datetime object), country and vehicle segment of desired footprint

    el_experiments = {'2020': {'year': 2020},
                       # '2021': {'year': start.year, 'start': start, 'end': end},  # example subannual period analysis
                       #  '2021fp': {'year': start.year, 'start':start, 'country':'FR', 'segment':'C'}  # example, single-country footprint (Figure 1)
                      }

    # define BEV experiments and parameter values (dict of dicts: {experiment_name: {experiment_params}})
    # accepted values for experiment_params:
    # BEV_life: int/float. Lifetime of BEV vehicles, in km
    # ICE_life: int/float. Lifetime of ICEV vehicles, in km
    # flowtrace: True/False. Assume flowtrace electricity trade assumption or grid average assumption (False)
    # allocation: True/False. Perform allocation to EOL of batteries where battery life (20% capacity fade) exceeds that of vehicle
    # energy_scen: str or False. Use different assumption for battery production electricity demand (False is default). str is experiment name, defined in Excel sheet.
    BEV_experiments = {'baseline': {'BEV_life':180000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':False},
                        'long_BEV_life': {'BEV_life':250000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':False},
                        'short_BEV_life': {'BEV_life':150000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':False},
                        'grid_avg': {'BEV_life':180000, 'ICE_life':180000, 'flowtrace':False, 'allocation':True, 'energy_scen':False},
                        'long_BEV_life_ga': {'BEV_life':250000, 'ICE_life':180000, 'flowtrace':False, 'allocation':True, 'energy_scen':False},
                        'short_BEV_life_ga': {'BEV_life':150000, 'ICE_life':180000, 'flowtrace':False, 'allocation':True, 'energy_scen':False},
                        'baseline_energy': {'BEV_life':180000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':'Gigafactory'},
                        'long_BEV_life_energy': {'BEV_life':250000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':'Gigafactory'},
                        'short_BEV_life_energy': {'BEV_life':150000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':'Gigafactory'},
                        'baseline_Ellingsen': {'BEV_life':180000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':'Ellingsen 2016'},
                        'baseline_Ellingsen_alloc': {'BEV_life':180000, 'ICE_life':180000, 'flowtrace':True, 'allocation':False, 'energy_scen':'Ellingsen 2016'},
                        'long_BEV_life_Ellingsen': {'BEV_life':250000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':'Ellingsen 2016'},
                        'short_BEV_life_Ellingsen': {'BEV_life':150000, 'ICE_life':180000, 'flowtrace':True, 'allocation':True, 'energy_scen':'Ellingsen 2016'},
                        'baseline_alloc': {'BEV_life':180000, 'ICE_life':180000, 'flowtrace':True, 'allocation':False, 'energy_scen':False},
                        'long_BEV_life_alloc': {'BEV_life':250000, 'ICE_life':180000, 'flowtrace':True, 'allocation':False, 'energy_scen':False},
                        'short_BEV_life_alloc': {'BEV_life':150000, 'ICE_life':180000, 'flowtrace':True, 'allocation':False, 'energy_scen':False},
                       }

    # this dict defines the groups of BEV_experiments to plot together in sensitivity analysis (Figure 6).
    # list items must therefore match BEV_experiment keys (experiment_name)
    sensitivity_groups = {'baseline':['baseline', 'long_BEV_life', 'short_BEV_life'],
                          'grid_average': ['grid_avg', 'long_BEV_life_ga', 'short_BEV_life_ga'],
                          'batt_energy': ['baseline_energy', 'long_BEV_life_energy', 'short_BEV_life_energy'],
                          'ellingsen': ['baseline_Ellingsen', 'long_BEV_life_Ellingsen', 'short_BEV_life_Ellingsen'],
						  'no_alloc': ['baseline_alloc', 'long_BEV_life_alloc', 'short_BEV_life_alloc']
						  }

    all_exp_results = {}  # holds all results for sensitivity analysis
    el_methodlist = []
    for exp, params in BEV_experiments.items():
        if BEV_experiments[exp]['flowtrace']:
            el_methodlist.append('flowtrace')
        elif not BEV_experiments[exp]['flowtrace']:
            el_methodlist.append('gridaverage')
    el_methodlist = list(set(el_methodlist))
    el_keys = list(itertools.product(el_experiments.keys(), el_methodlist))
    el_figs = {key: True for key in el_keys}  # keeps track if electricity figures have already been plotted for the electriicty scenario


    # Begin running electricity experiments
    # Default is to recalculate everything
    for el_exp, params in el_experiments.items():
        logging.info(f'Starting el_experiment {el_exp}')

        if ('year' not in params.keys()) and ('start' in params.keys()):
            # manually populate the 'year' entry manually if missing
            params['year'] = params[start].year

        if set(params.keys()) == exp_type_dict['fullyear']:
            experiment_type = 'fullyear'
        elif set(params.keys()) == exp_type_dict['subyear']:
            experiment_type = 'subyear'
        elif set(params.keys()) == exp_type_dict['country_fp']:
            if params['segment'] in segments:
                experiment_type = 'country_fp'
                if params['segment'] in segment_dict.keys():
                    params['segment'] = segment_dict[params['segment']]  # convert to letter segment classification
            else:
                logging.error(f'Incorrect specification of desired vehicle segment')
        else:
            logging.error('Invalid electricity experiment input')

        # case for sub-annual analysis period (or single-country footprint)
        if (experiment_type == 'subyear') or (experiment_type == 'country_fp'):
            if new_query:
                bentso_extract.bentso_query(params['year'])  # can alternatively specify year of analysis here
            logging.info(f'year: {params["year"]}')
            if experiment_type == 'subyear':
                # for sub-annual analysis period
                logging.info(f'subperiod start: {params["start"]}')
                logging.info(f'subperiod end: {params["end"]}')
                ei_countries = entso_data_clean.clean_entso(year=params['year'],
                                                            start=params['start'],
                                                            end=params['end'])
            else:
                # For country-footprint experiment (snapshot of time and segment)
                # identical to subannual analysis; only visualization changes
                if new_query:
                    # NB: we can use full_year = True here, but query will be rounded to closest hour
                    # (and which is averaged out if sub-hourly samples are available)
                    bentso_extract.bentso_query(params['year'], full_year=False)
                logging.info(f'country footprint: {params["country"]}')
                logging.info(f'segment footprint: {params["segment"]}')
                ei_countries, timestamp = entso_data_clean.clean_entso(year=params['year'],
                                                                        start=params['start'],
                                                                        end=params['start'] + timedelta(hours=1),
                                                                        country=params['country'])

                if np.abs(timestamp.to_pydatetime() - params['start']) > timedelta(hours=12):
                    print('Warning: Large discrepancy between sampled time and user defined time. Try running a new query.')
                    print('Completing analysis on closest time available')

        # case for full-year analysis
        elif experiment_type == 'fullyear':
            if new_query:
                bentso_extract.bentso_query(params['year'])  # can alternatively specify year of analysis here
            logging.info(f'year: {params["year"]}')
            ei_countries = entso_data_clean.clean_entso(year=params['year'])

        print(f'Using ecoinvent factors for {ei_countries} \n')
        # Calculate hybridized impact factors for electricity
        logging.info('Starting calculation of hybridized emission factors \n')
        # ef_countries, no_ef, countries_missing_ef = hybrid.hybrid_emission_factors(ei_countries, params['year'])
        countries_missing_ef = hybrid.clean_impact_factors(params['year'], ei_countries, calc_hybrid)


        # Begin BEV experiments - calculation and visualization
        for BEV_exp, BEV_params in BEV_experiments.items():
            print('\n **********')
            print(f'Performing experiment {BEV_exp} with electricity experiment {el_exp}')
            logging.info(f'Starting BEV experiment {BEV_exp}')

            if BEV_params['flowtrace']:
                el_method = 'flowtrace'
            elif not BEV_params['flowtrace']:
                el_method = 'gridaverage'
            all_exp_results[BEV_exp], ICEV_prodEOL_impacts, ICEV_op_int, SI_fp_temp = calc.run_calcs(run_id=BEV_exp+'_'+el_exp,
                                                      year = params['year'],
                                                      no_ef_countries = countries_missing_ef,
                                                      BEV_lifetime=BEV_params['BEV_life'],
                                                       ICEV_lifetime=BEV_params['ICE_life'],
                                                      flowtrace_el=BEV_params['flowtrace'],
													  allocation = BEV_params['allocation'],
                                                      production_el_intensity=679,
                                                      export_data=True,
                                                      include_TD_losses=True,
                                                      incl_ei=False,
                                                      energy_sens=BEV_params['energy_scen']
                                                      )
            if BEV_exp == 'baseline':
                SI_fp = SI_fp_temp  # only need to update the main SI file (that is submitted with article)
            logging.info('Visualizing results...')
            if (experiment_type == 'fullyear') or (experiment_type == 'subyear'):
                viz.visualize(BEV_exp+'_'+el_exp, export_figures=True, include_TD_losses=True, plot_el=el_figs[(el_exp, el_method)])
                if el_figs[(el_exp, el_method)]:
                    el_figs[(el_exp, el_method)] = False
            elif experiment_type == 'country_fp':
                viz.country_footprint(BEV_exp+'_'+el_exp, params, timestamp, export_figures=True)

            logging.info(f'Completed experiment electricity {el_exp}, BEV scenario {BEV_exp}')

        # Check we have all the necessary experiments to run a sensitivity analysis
        run_sensitivity = False
        for sens, exp in sensitivity_groups.items():
            if len(set(exp) - set(exp).intersection(set(BEV_experiments.keys()))) == 0:
                run_sensitivity = True

        # Compile BEV intensities from all experiments for sensitivity analysis
        if run_sensitivity and (experiment_type != 'country_fp'):
            mi = pd.MultiIndex.from_product([all_exp_results.keys(), all_exp_results['baseline'].columns.tolist()])
            results = pd.DataFrame(index=all_exp_results['baseline'].index, columns=mi)  # countries x all experiments

            for exp in all_exp_results.keys():
                results[exp] = all_exp_results[exp]

            # Compile ICEV intensities for different lifetimes; these are constant for all scenarios
            tmp_ICEV_fp = ICEV_prodEOL_impacts.to_frame().T.reindex(results.index)
            ind = pd.MultiIndex.from_product([['ICEV 250k', 'ICEV 200k', 'ICEV 180k'], tmp_ICEV_fp.columns.tolist()])
            tmp_ICEV_fp = tmp_ICEV_fp.reindex(ind, axis=1)

            ICEV_dict = {'ICEV 250k': 250000,
                         'ICEV 200k': 200000,
                         'ICEV 180k': 180000}

            for ICE, dist in ICEV_dict.items():
                for seg in tmp_ICEV_fp[ICE].columns:
                    tmp_ICEV_fp.loc[:, (ICE, seg)] = ((ICEV_prodEOL_impacts / dist * 1e6).add(ICEV_op_int, axis=0)).loc[seg]

            # concatenate all results and visualize sensitivity analysis
            results = pd.concat([results, tmp_ICEV_fp], axis=1)
            results = results.loc[~results.index.isin(ei_countries)]
            viz.plot_fig6(results, sensitivity_groups, export_figures=True)

            results = results.reindex(columns=list(BEV_experiments.keys()) + ['ICEV 205k', 'ICEV 200k', 'ICEV 180k'], level=0)

            #-- Begin export to SI
            # separate each 'set' of BEV_experiments for exporting to SI
            results_lifetime_sensitivity = results.iloc[:, np.r_[0:12, -12:0]].copy()
            if results.shape[1] > 24:
                results_ga = results.iloc[:, np.r_[12:24]].copy()

            # rename columns for human-readability
            new_ind = pd.MultiIndex.from_product([['Baseline (180k)', 'Long BEV life (250k)', 'Short BEV life (150k)', 'ICEV 250k','ICEV 200k','ICEV 180k'], ['A','C','JC','JE']])

            results_lifetime_sensitivity.columns = new_ind
            results_ga.columns = new_ind[0:12]

            # make custom dataframe for low-manufacturing-energy sensitivity
            df = (all_exp_results['baseline_energy'] - all_exp_results['baseline']) / all_exp_results['baseline']

            sens_batt_energy = pd.concat([all_exp_results['baseline_energy'].round(0), df], axis=1)
            sens_batt_energy.columns = pd.MultiIndex.from_product([['Footprint g CO2e/km', '% change from baseline'], df.columns])

            sens_ellingsen = pd.concat([all_exp_results['baseline_Ellingsen'].round(0), df], axis=1)
            sens_ellingsen.columns = pd.MultiIndex.from_product([['Footprint g CO2e/km', '% change from baseline'], df.columns])

            # dict of captions for each sheet
            excel_dict2 = {'Table S12': 'Sensitivity with alternative battery production electricity. Footprint with lower electricity demand and % change from baseline',
                           'Table S13': 'Data for Figure 7. Sensitivity with differing vehicle lifetimes for BEVs and ICEVs. Lifecycle carbon intensity, in g CO2e/vkm.',
                           'Table S14': 'BEV footprints using grid-average approach from Moro and Lonza (2018) to calculate electricity mix footprint, in g CO2e/vkm',
                           'Table S15': 'Sensitivity with Ellingsen 2016 battery production electricity values. Footprint with lower electricity demand and % change from baseline'}

            book = openpyxl.load_workbook(SI_fp)
            with pd.ExcelWriter(SI_fp, engine="openpyxl") as writer:
                writer.book = book
                sens_batt_energy.to_excel(writer, 'Table S12', startrow=2)
                results_lifetime_sensitivity.round(0).to_excel(writer, 'Table S13', startrow=2)
                results_ga.round(0).to_excel(writer, 'Table S14', startrow=2)
                sens_ellingsen.to_excel(writer, 'Table S15', startrow=2)

                for sheet, caption in excel_dict2.items():
                    worksheet = writer.sheets[sheet]
                    worksheet['A1'] = sheet
                    worksheet['A1'].font = Font(color='0070C0', bold=True)

                    worksheet.merge_cells('C1:L1')
                    worksheet['C1'] = caption
                    worksheet['C1'].alignment = Alignment(wrap_text=True)

                    worksheet.row_dimensions[1].height = 60

                    if sheet == 'Table S12' or sheet == 'Table S15':
                        for col in ['F','G','H','I']:
                            for row in range(worksheet.min_row, worksheet.max_row + 1):
                                worksheet[col + str(row)].number_format = '#0%'

                writer.book.save(SI_fp)